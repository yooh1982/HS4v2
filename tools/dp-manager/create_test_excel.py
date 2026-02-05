#!/usr/bin/env python3
"""
테스트용 엑셀 파일 생성 스크립트
사용법: python3 create_test_excel.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import sys

def create_test_excel(filename="H2567_IMO9991862_IOList_20260125.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "IOLIST"
    
    # 헤더 행
    headers = ["IO_NO", "IO_NAME", "IO_TYPE", "DESCRIPTION", "REMARKS"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # 샘플 데이터
    sample_data = [
        ["IO001", "Main Engine Start", "Digital Input", "Main engine start signal", "Critical"],
        ["IO002", "Main Engine Stop", "Digital Input", "Main engine stop signal", "Critical"],
        ["IO003", "Thruster Control", "Analog Output", "Thruster control signal", "Important"],
        ["IO004", "DP System Status", "Digital Output", "DP system status indicator", "Normal"],
        ["IO005", "Power Management", "Digital Input", "Power management signal", "Important"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    
    wb.save(filename)
    print(f"✅ 테스트 엑셀 파일 생성 완료: {filename}")
    print(f"   파일명에서 Hull NO: H2567, IMO: IMO9991862 자동 추출됩니다.")

if __name__ == "__main__":
    filename = sys.argv[1] if len(sys.argv) > 1 else "H2567_IMO9991862_IOList_20260125.xlsx"
    create_test_excel(filename)
