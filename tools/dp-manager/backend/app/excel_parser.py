import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger("dp-manager")

# Protocol별 필수 컬럼 정의
# Protocol: MQTT, NMEA, OPCUA, OPCDA, MODBUS
PROTOCOL_COLUMNS = {
    "MQTT": {
        "required_columns": [
            "Resource",
            "Data type",
            "RuleNaming",
            "Level 1",
            "Level 2",
            "Level 3",
            "Level 4",
            "Miscellaneous",
            "Measure",
            "Description",
            "Calculation",
            "MQTT Tag",
            "Remark"
        ],
        "required_value_columns": [
            "Resource",
            "Data type",
            "RuleNaming",
            "Level 1",
            "Measure",
            "MQTT Tag"
        ]
    },
    # 다른 프로토콜은 나중에 추가
    # "NMEA": {...},
    # "OPCUA": {...},
    # "OPCDA": {...},
    # "MODBUS": {...},
}

# 기본값은 MQTT (하위 호환성)
REQUIRED_COLUMNS = PROTOCOL_COLUMNS["MQTT"]["required_columns"]
REQUIRED_VALUE_COLUMNS = PROTOCOL_COLUMNS["MQTT"]["required_value_columns"]


def get_protocol_columns(protocol: str = "MQTT") -> tuple[List[str], List[str]]:
    """
    Protocol별 필수 컬럼 반환
    
    Returns:
        (required_columns, required_value_columns)
    """
    protocol_upper = protocol.upper()
    if protocol_upper in PROTOCOL_COLUMNS:
        cols = PROTOCOL_COLUMNS[protocol_upper]
        return cols["required_columns"], cols["required_value_columns"]
    else:
        # 기본값은 MQTT
        logger.warning(f"알 수 없는 프로토콜: {protocol}. MQTT 컬럼을 사용합니다.")
        return REQUIRED_COLUMNS, REQUIRED_VALUE_COLUMNS


def validate_required_columns(headers: List[str], protocol: str = "MQTT") -> tuple[bool, List[str]]:
    """
    Protocol별 필수 컬럼이 모두 있는지 검증
    
    Args:
        headers: 엑셀 헤더 리스트
        protocol: Protocol 타입 (MQTT, NMEA, OPCUA, OPCDA, MODBUS)
    
    Returns:
        (is_valid, missing_columns)
    """
    header_set = {str(h).strip() for h in headers if h}
    required_columns, _ = get_protocol_columns(protocol)
    required_set = set(required_columns)
    missing = list(required_set - header_set)
    return len(missing) == 0, missing


def parse_iolist_excel(file_path: str, protocol: str = "MQTT") -> List[Dict[str, Any]]:
    """
    IOLIST 엑셀 파일의 "IOList" 시트를 파싱하여 리스트로 반환
    
    Args:
        file_path: 엑셀 파일 경로
        protocol: Protocol 타입 (MQTT, NMEA, OPCUA, OPCDA, MODBUS). 기본값은 MQTT
    
    엑셀 구조:
    - "IOList" 시트: 첫 번째 행은 헤더, 두 번째 행부터 데이터
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        
        # "IOList" 시트 찾기
        if "IOList" not in wb.sheetnames:
            raise ValueError("엑셀 파일에 'IOList' 시트가 없습니다.")
        
        ws = wb["IOList"]
        
        items = []
        headers = []
        
        # 첫 번째 행에서 헤더 읽기
        for cell in ws[1]:
            headers.append(cell.value if cell.value else "")
        
        # 헤더가 없으면 에러
        if not headers or all(not h for h in headers):
            raise ValueError("엑셀 파일에 헤더가 없습니다.")
        
        # Protocol별 필수 컬럼 검증
        is_valid, missing_columns = validate_required_columns(headers, protocol)
        if not is_valid:
            required_columns, _ = get_protocol_columns(protocol)
            raise ValueError(
                f"필수 컬럼이 누락되었습니다: {', '.join(missing_columns)}. "
                f"Protocol: {protocol}, 필수 컬럼: {', '.join(required_columns)}"
            )
        
        logger.info(f"필수 컬럼 검증 완료. 헤더 수: {len([h for h in headers if h])}")
        
        # 헤더를 정확한 이름으로 매핑 (대소문자 구분)
        header_map = {}
        for idx, header in enumerate(headers):
            if header:
                header_map[idx] = str(header).strip()
        
        # 두 번째 행부터 데이터 읽기
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # 빈 행은 건너뛰기
            if all(cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == "") for cell in row):
                continue
            
            item = {}
            has_data = False
            
            for col_idx, cell in enumerate(row):
                if col_idx in header_map:
                    header_name = header_map[col_idx]
                    value = cell.value
                    
                    # 값 처리
                    if value is not None:
                        if isinstance(value, (int, float)):
                            # 숫자는 문자열로 변환
                            value = str(value)
                        else:
                            value = str(value).strip() if value else None
                        
                        if value:
                            has_data = True
                    else:
                        value = None
                    
                    item[header_name] = value
            
            # 데이터가 있는 행만 추가
            if has_data:
                items.append(item)
        
        logger.info(f"엑셀 파일에서 {len(items)}개의 항목을 파싱했습니다.")
        return items
        
    except Exception as e:
        logger.error(f"엑셀 파일 파싱 중 오류 발생: {e}", exc_info=True)
        raise


def parse_device_sheet(file_path: str) -> Dict[str, str]:
    """
    엑셀 파일의 "Device" 시트를 파싱하여 Device별 Protocol 정보를 반환
    
    엑셀 구조:
    - "Device" 시트: 첫 번째 행은 헤더 (Device, Protocol 등)
    - 두 번째 행부터: Device별 Protocol 정보
    
    Returns:
        Dict[device_name, protocol]: Device 이름을 키로, Protocol을 값으로 하는 딕셔너리
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        
        # "Device" 시트 찾기
        if "Device" not in wb.sheetnames:
            logger.warning("엑셀 파일에 'Device' 시트가 없습니다. 빈 딕셔너리를 반환합니다.")
            return {}
        
        ws = wb["Device"]
        
        # 첫 번째 행에서 헤더 읽기
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        # Device와 Protocol 컬럼 찾기
        device_col_idx = None
        protocol_col_idx = None
        
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if "device" in header_lower and device_col_idx is None:
                device_col_idx = idx
            elif "protocol" in header_lower and protocol_col_idx is None:
                protocol_col_idx = idx
        
        if device_col_idx is None:
            logger.warning("'Device' 시트에 'Device' 컬럼을 찾을 수 없습니다.")
            return {}
        
        if protocol_col_idx is None:
            logger.warning("'Device' 시트에 'Protocol' 컬럼을 찾을 수 없습니다.")
            return {}
        
        # 두 번째 행부터 데이터 읽기
        device_protocol_map = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            device_cell = row[device_col_idx]
            protocol_cell = row[protocol_col_idx] if protocol_col_idx < len(row) else None
            
            device_name = None
            protocol = None
            
            if device_cell and device_cell.value:
                device_name = str(device_cell.value).strip()
            
            if protocol_cell and protocol_cell.value:
                protocol = str(protocol_cell.value).strip()
            
            if device_name:
                device_protocol_map[device_name] = protocol or "MQTT"  # 기본값은 MQTT
        
        logger.info(f"Device 시트에서 {len(device_protocol_map)}개의 Device를 파싱했습니다.")
        return device_protocol_map
        
    except Exception as e:
        logger.error(f"Device 시트 파싱 중 오류 발생: {e}", exc_info=True)
        # Device 시트 파싱 실패는 경고만 하고 계속 진행
        return {}


def generate_data_channel_id(item: Dict[str, Any]) -> str:
    """
    DataChannelId 생성: "/RuleNaming/Level1/Level2/Level3/Level4/Miscellaneous/Measure"
    """
    parts = [
        item.get("RuleNaming", ""),
        item.get("Level 1", ""),
        item.get("Level 2", ""),
        item.get("Level 3", ""),
        item.get("Level 4", ""),
        item.get("Miscellaneous", ""),
        item.get("Measure", "")
    ]
    # 빈 값은 제외하고 조합
    filtered_parts = [str(p).strip() for p in parts if p and str(p).strip()]
    return "/" + "/".join(filtered_parts)


def extract_iolist_fields(item: Dict[str, Any], validate_required_values: bool = True) -> Dict[str, Any]:
    """
    MQTT 프로토콜 파싱된 딕셔너리에서 IOLIST 필드를 추출
    
    필수 컬럼을 검증하고, 모든 필드를 extra_data에 저장
    주요 필드는 별도 필드에도 저장 (호환성)
    
    필수 컬럼은 헤더에 있어야 하지만, 값은 null일 수 있음
    단, 필수 값 컬럼은 반드시 값이 있어야 함
    """
    import json
    
    # 필수 컬럼이 딕셔너리에 키로 존재하는지 검증 (값은 null 가능)
    missing_required = []
    for col in REQUIRED_COLUMNS:
        if col not in item:
            missing_required.append(col)
    
    if missing_required:
        raise ValueError(
            f"필수 컬럼이 누락되었습니다: {', '.join(missing_required)}"
        )
    
    # 필수 값 검증 (validate_required_values가 True일 때만)
    if validate_required_values:
        missing_values = []
        for col in REQUIRED_VALUE_COLUMNS:
            value = item.get(col)
            if not value or (isinstance(value, str) and not value.strip()):
                missing_values.append(col)
        
        if missing_values:
            raise ValueError(
                f"필수 값이 누락되었습니다: {', '.join(missing_values)}"
            )
    
    # DataChannelId 생성
    data_channel_id = generate_data_channel_id(item)
    
    # 주요 필드 추출 (호환성을 위해)
    result = {
        "io_no": item.get("MQTT Tag", ""),  # MQTT Tag를 io_no로 사용
        "io_name": item.get("Description", "") or item.get("Measure", ""),
        "io_type": item.get("Data type", ""),
        "description": item.get("Description", ""),
        "remarks": item.get("Remark", ""),
        "extra_data": None
    }
    
    # 모든 필드를 extra_data에 JSON으로 저장 (DataChannelId 포함)
    all_data = {}
    for key, value in item.items():
        if value is not None:
            if isinstance(value, str):
                value = value.strip()
            all_data[key] = value
    
    # DataChannelId 추가
    all_data["DataChannelId"] = data_channel_id
    
    result["extra_data"] = json.dumps(all_data, ensure_ascii=False)
    
    return result
